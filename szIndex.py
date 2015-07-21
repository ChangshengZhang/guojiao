#!/usr/bin/python
# -*- coding: utf-8 -*-
# File Name: szIndex.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

#########################################################################

# 需要用到 openpyxl 模块读取 excel 文件
from openpyxl.reader.excel import load_workbook
#import numpy as np
import matplotlib.pyplot as plt

# 读取excel文件信息
# 将最高价、最低价、收盘价等信息存入到数组
def load_data(filename):
    #载入excel文件数据
    # ws得到的是文件指针
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # 下面5个数组依次是
    # 交易时间、最高价、最低价、交易量、收盘价
    stock_date = []
    price_high = []
    price_low = []
    stock_volume = []
    price_closed = []
    # 循环读取每一行的数据
    # openpyxl 用 cell(row = , column = ).value 函数得到某个单元格的值
    empty_str=  ws.cell(row=1,column=1).value
    for rx in range(5,ws.get_highest_row()+1):
        if ws.cell(row=rx,column=2).value!=empty_str and ws.cell(row=rx,column=3).value!=empty_str:
            date = str(ws.cell(row =rx ,column = 1).value)
            stock_date.append(date.split(" ")[0])
            price_high.append(ws.cell(row=rx,column=2).value)
            price_low.append(ws.cell(row=rx,column=3).value)
            stock_volume.append(ws.cell(row=rx,column=4).value)
            price_closed.append(ws.cell(row=rx,column=5).value)

    return stock_date,price_high,price_low,stock_volume,price_closed

# 得到一级和二级高点
# 一级高点是指 前后个两个交易日的最高价比当日的最高价低
# 二级高点是指 前面两个和后面一个一级高点的价格比该一级高点的最高价低
def get_extreme_high_point(price_high,compared_day):
    # compared_day 是一级高点比较的时间长度，可以自定义，目前设置的是两天

    #数组成对使用，XX 数组是价格，另一个数组 XX_index 是交易日的索引，即第几个交易日
    # 一级高点的有关数组
    extreme_high_point = []
    extreme_high_point_index = []
    price_high_new= []
    # 二级高点的有关数组
    high_point =[]
    high_point_index = []
    # 卖点的有关数组
    sell_point = []
    sell_point_index = []
    #构造一个新数组，在数组最初和最尾各自新增几个数字，避免比较的时候数组长度溢出的问题
    for i in range(compared_day):
        price_high_new.append(price_high[0])

    for i in range(len(price_high)):
        price_high_new.append(price_high[i])

    for i in range(compared_day):
        price_high_new.append(price_high[len(price_high)-1])

    # 寻找一级高点
    # flag是标志符，如果符合一级高点的定义，则 flag设置为1，不符合则为0
    for i in range(len(price_high)):
        flag =1
        for j in range(compared_day*2+1):
            if price_high_new[i+compared_day]< price_high_new[i+j]:
                flag = 0
                break
        if flag == 1:
            extreme_high_point_index.append(i)
            extreme_high_point.append(price_high[i])
    # 寻找二级高点
    # 利用一级高点的数据
    # flag的用法和寻找一级高点一样
    # 注意循环的长度，最后一个一级高点是要舍去的，因为此时尚未出现卖点
    for i in range(len(extreme_high_point)-1):
        flag =1
        # 忽略最先的交易日，避免数组长度溢出
        if i < compared_day:
            continue
        else:
            for j in range(compared_day):
                if extreme_high_point[i-j-1]>extreme_high_point[i]:
                    flag=1
                    break
                else:
                    flag=flag+1
                if flag == compared_day+1:
                    if extreme_high_point[i+1]<extreme_high_point[i]:
                        high_point.append(extreme_high_point[i])
                        high_point_index.append(extreme_high_point_index[i])
                        sell_point.append(extreme_high_point[i+1])
                        sell_point_index.append(extreme_high_point_index[i+1])

    # 返回二级高点和卖点
    return high_point,high_point_index,sell_point,sell_point_index,extreme_high_point,extreme_high_point_index

    #return extreme_high_point,extreme_high_point_index

# 得到一级和二级低点
# 思路和寻找一级和二级高点相似
# 函数的构造过程参加 get_extreme_high_point()
def get_extreme_low_point(price_low,compared_day):
    extreme_low_point = []
    extreme_low_point_index = []
    price_low_new= []
    low_point =[]
    low_point_index =[]
    buy_point =[]
    buy_point_index =[]
    #新数组
    for i in range(compared_day):
        price_low_new.append(price_low[0])

    for i in range(len(price_low)):
        price_low_new.append(price_low[i])

    for i in range(compared_day):
        price_low_new.append(price_low[len(price_low)-1])


    for i in range(len(price_low)):
        flag =1
        for j in range(compared_day*2+1):
            if price_low_new[i+compared_day]> price_low_new[i+j]:
                flag = 0
                break
        if flag == 1:
            extreme_low_point_index.append(i)
            extreme_low_point.append(price_low[i])

    for i in range(len(extreme_low_point)-1):
        flag =1
        if i < compared_day:
             continue
        else:
            for j in range(compared_day):
                if extreme_low_point[i-j-1]<extreme_low_point[i]:
                    flag =1
                    break
                else:
                    flag =flag+1
                if flag ==compared_day+1:
                    #反弹条件
                    if extreme_low_point[i+1]>extreme_low_point[i]:
                        low_point.append(extreme_low_point[i])
                        low_point_index.append(extreme_low_point_index[i])
                        buy_point.append(extreme_low_point[i+1])
                        buy_point_index.append(extreme_low_point_index[i+1])

    return low_point,low_point_index,buy_point,buy_point_index
    #return extreme_low_point,extreme_low_point_index

# 合并两个数组
# 因为卖点和买点是两个数组，故需要合并
# 合并的思路是，按照交易日的大小合并
# 算法和“两个有序数组合并”相同
def mix_signal_array(high_point_index,low_point_index):
    # 两个索引
    temp_high_index = 0
    temp_low_index  = 0
    # 合并后的数组，第一个是  合并后的交易日索引，第二个是 合并后的 交易信号：buy（b） or sell（s）
    mix_index = []
    mix_signal = []
    # 合并算法
    for i in range(len(high_point_index)+len(low_point_index)):

        if temp_high_index<len(high_point_index) and temp_low_index<len(low_point_index):
            if  high_point_index[temp_high_index]<low_point_index[temp_low_index]:
                mix_index.append(high_point_index[temp_high_index])
                temp_high_index = temp_high_index+1
                mix_signal.append('s')
            elif high_point_index[temp_high_index]>low_point_index[temp_low_index]:
                mix_index.append(low_point_index[temp_low_index])
                temp_low_index = temp_low_index+1
                mix_signal.append('b')
            else:
                # 如果同一天两个指标同时出现，则舍弃
                temp_low_index = temp_low_index+1
                temp_high_index = temp_high_index+1
                i = i+1
        elif temp_high_index ==len(high_point_index) and temp_low_index<len(low_point_index):
            mix_index.append(low_point_index[temp_low_index])
            mix_signal.append('b')
            temp_low_index = temp_low_index+1
        elif temp_high_index <len(high_point_index) and temp_low_index==len(low_point_index):
             mix_index.append(high_point_index[temp_high_index])
             temp_high_index = temp_high_index+1
             mix_signal.append('s')

    return mix_index,mix_signal

#计算收益率
def cal_revenue(price_closed,mix_index,mix_signal,trade_discounter =1):

    # 将交易信号进一步处理
    # 因为并非所有的交易信号都需要执行
    # 交易策略可以自定义
    # 目前的交易策略是：
    #   1. 先买后卖，不能做空
    #   2. 本金为100，只能做一笔买进，之前的仓位不平掉，不能再买进
    #   3. 交易费用为 0.1%，买进卖出都要计算

    # flag 是标志位，取0,1值，依次来判断当前的仓位情况：空仓或满仓
    flag =0
    # new_index, new_signal 的值分别是需要买进或者卖出的日期及交易行为（买或卖）
    new_index = []
    new_signal = []
    for i in range(len(mix_index)):
        if mix_signal[i]=='b':
            if flag==0:
                flag=1
                new_index.append(mix_index[i])
                new_signal.append(mix_signal[i])
        else:
            if flag==1:
                new_index.append(mix_index[i])
                new_signal.append(mix_signal[i])
                flag =0

    # flag的用法同上
    flag =0
    revenue = []
    action_index =0
    #base 是上一日的资产
    base =100
    for i in range(len(price_closed)):
        #前一天是空仓
        if flag ==0:
            # 如果有买进信号的话，按照当天收盘价买入
            # 否则一直空仓
            if i ==new_index[action_index] and new_signal[action_index] =='b':
                base = base*trade_discounter
                revenue.append(base)
                flag =1
                action_index =action_index+1
                if action_index == len(new_index):
                    action_index = action_index-1
            else:

                revenue.append(base)

        #前一天有仓位
        else:
            # 如果有卖出信号，则按照当天收盘价卖出
            # 否则一直持有
            if i == new_index[action_index] and new_signal[action_index]=='s':
                base = base*price_closed[i]/price_closed[i-1]
                flag=0
                action_index =action_index+1
                if action_index == len(new_index):
                    action_index =action_index-1
                base =base*trade_discounter
                revenue.append(base)
            #继续持仓
            else:
                base = base*price_closed[i]/price_closed[i-1]
                revenue.append(base)



    return revenue,new_index,new_signal

# 计算年化收益率
def cal_annual_rev(revenue,stock_date):

    # 寻找每一年的起始日
    start_year_index = []
    last_year = 0
    annual_rev =[]
    year_number=[]
    for i in range(len(stock_date)):
        temp_date = str(stock_date[i])
        temp_year = temp_date.split("-")

        if i ==0:
            last_year =temp_year[0]
            start_year_index.append(i)
            year_number.append(last_year)
        elif i == len(stock_date)-1:
            start_year_index.append(i)
        else:
            if temp_year[0] != last_year:
                last_year =temp_year[0]
                start_year_index.append(i)
                year_number.append(temp_year[0])


    for i in range(len(start_year_index)-1):
        annual_rev.append(revenue[start_year_index[i+1]]/revenue[start_year_index[i]]-1)


    return annual_rev,year_number


# 快速下跌后的反弹收益情况
def cal_bounce_after_fall(filename,high_point_index,price_high,price_low,stock_date):
    bounce = []
    temp_date = []
    temp_filename =-1
    # 上证只取2000年之后的数据
    if filename=="sz.xlsx":
        temp_filename =2000
    for i in range(len(high_point_index)-1):
        min = 1000000
        date = 0

        if high_point_index[i]>temp_filename:
            for j in range(high_point_index[i],high_point_index[i+1]+1):
                if min >price_low[j]:
                    min = price_low[j]
                    date =j


            if price_high[high_point_index[i]]/min >1.25:
                a= ("%.1f" % float((1-min/price_high[high_point_index[i]])*100))
                b = ("%.1f" % float((price_high[high_point_index[i+1]]/min-1)*100))

                #第一个高点日期，点数，最低点日期，点数，下跌天数，下跌幅度，第二个高点日期，点数，上涨天数，上涨幅度
                temp_date.append([stock_date[high_point_index[i]],int(price_high[high_point_index[i]]),stock_date[date],int(min),date-high_point_index[i],
                                  a,stock_date[high_point_index[i+1]],int(price_high[high_point_index[i+1]]),high_point_index[i+1]-date,b])

                bounce.append(price_high[high_point_index[i+1]]/min-1)



    return bounce,temp_date



if __name__  == "__main__":
    filename = "sz.xlsx"
    #比较的时间长度
    compared_day = 5
    stock_date,price_high,price_low,stock_volume,price_closed = load_data(filename)


    #调用函数，得到高点和低点
    extreme_high_point,extreme_high_point_index,sell_point,sell_point_index,high_point,high_point_index = get_extreme_high_point(price_high,compared_day)

    extreme_low_point, extreme_low_point_index,buy_point,buy_point_index = get_extreme_low_point(price_low,compared_day)

    # 得到混合后的数组
    mix_index,mix_signal = mix_signal_array(sell_point_index,buy_point_index)

    #计算收益，分别计算不考虑交易费用和考虑交易费用的收益
    revenue,new_index,new_signal = cal_revenue(price_closed,mix_index,mix_signal)
    revenue_discounter,new_index_discounter,new_signal_discounter = cal_revenue(price_closed,mix_index,mix_signal,0.999)

    #计算年华收益率，分别计算不考虑交易费用和不考虑交易费用
    annual_rev,annual_year_number = cal_annual_rev(revenue,stock_date)
    annual_rev_discounter,annual_year_number_discounter = cal_annual_rev(revenue_discounter,stock_date)


    bounce,temp_data = cal_bounce_after_fall(filename,high_point_index,price_high,price_low,stock_date)

    # output =open('data.txt','w')
    # output.write("High Point:\n \n")
    # for i in range(len(extreme_high_point)):
    #     output.write(str(extreme_high_point_index[i]))
    #     output.write("\n")
    # output.write("\n Low Point:\n\n")
    # for i in range(len(extreme_low_point_index)):
    #     output.write(str(extreme_low_point_index[i]))
    #     output.write("\n")
    #
    # output.close()
    # 输出到文件


    # output_filename = filename+"_data.txt"
    # output = open(output_filename,'w')
    # str_title = "高点一日期 & 点数 & 最低点日期 & 点数 & $\downarrow$ 天 &$\downarrow$幅度 & 高点二日期 & 点数 & $\uparrow$天& $\uparrow$ 幅度 \\\ \hline"
    # output.write(str_title)
    # output.write("\n")
    # a =0
    # b =0
    # c= 0
    # d=0
    # for i in range(len(temp_data)):
    #     for j in range(len(temp_data[0])):
    #         if j == 4:
    #             a =a + float(temp_data[i][j])
    #         if j ==5:
    #             b = b+ float(temp_data[i][j])
    #         if j ==8:
    #             c = c+ float(temp_data[i][j])
    #         if j ==9:
    #             d =d+ float(temp_data[i][j])
    #
    #         if j ==5:
    #             output.write("\\textcolor{blue}{")
    #             output.write(str(temp_data[i][j]))
    #             output.write("\%")
    #             output.write("}")
    #         elif j==9:
    #             output.write("\\textcolor{red}{")
    #             output.write(str(temp_data[i][j]))
    #             output.write("\%")
    #             output.write("}")
    #         else:
    #             output.write(str(temp_data[i][j]))
    #
    #         if j != len(temp_data[0])-1:
    #             output.write(" & ")
    #     output.write("  \\\  \hline")
    #     output.write("\n")
    # output.close()
    #
    # print  a/len(temp_data),b/len(temp_data),c/len(temp_data),d/len(temp_data)

    # 画图，需要用到 matplotlib 这个库
    # plt.plot(price_high)
    # plt.plot(extreme_high_point_index,extreme_high_point,"r+",linewidth=5)
    # plt.plot(price_low,"g")
    # plt.plot(extreme_low_point_index,extreme_low_point,"b+",linewidth=5)
    #


    # plt.plot(revenue,"b",label = "No trading fee")
    # plt.plot(revenue_discounter,"r", label = "Trading fee is 0.1%")


    #plt.plot(bounce)
    plt.plot(annual_year_number,annual_rev,"b",label = "No trading fee")
    plt.plot(annual_year_number_discounter,annual_rev_discounter,"r",label = "Trading fee is 0.1%")

    plt.legend(loc='upper left')
    plt.grid(True)
    plt.xlabel("Time(day)")
    plt.ylabel("Point")
    plt.title("Revenue")
    plt.show()